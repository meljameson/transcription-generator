'''
This script fills in the given .txt template with values from a .xlsx file
It outputs the result to a new file which is named after the file_naming_convention variable
The file_naming_convention is a column header in the excel file, such as "formNumber"

Example:
Template: "In honor of `petitionerName`, on this day `date`" 
Result: "In honor of Abraham Lincoln, on this day 1/2/1803"
In this case, the excel file must have columns named petitionerName and date

'''
# -*- coding: utf-8 -*-

# imports for reading excel docs
import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string

class Transcriber:
		
	def __init__(self, excel_name, template_name, save_name, file_name, type, append_name):
		self.excel_doc = excel_name
		self.template = template_name
		self.file_save_location = save_name
		self.file_naming_convention = file_name
		self.file_type = type
		self.file_to_append = save_name + append_name
		self.worksheet = self.get_worksheet()
		
	def toString(self):
		print(self.excel_doc)
		print(self.template)
		print(self.file_save_location)
		print(self.file_naming_convention)
		print(self.file_type)
		print(self.file_to_append)
		
	'''
	# identifies the excel document to use data from
	excel_doc

	# identifies the template to fill out
	template

	# identifies what folder you want the transcriptions to be saved in
	file_save_location

	# name the file after a column in the excel file
	file_naming_convention = ""

	# what file format to save the file as
	file_type = ""

	# if appending, change this
	file_to_append = file_save_location + ""
	'''
	
	# get workbook
	def get_worksheet(self):
		workbook = openpyxl.load_workbook(self.excel_doc)
		# get first worksheet. can be replaced by workbook.get_sheet_by_name('Sheet1')
		self.worksheet = workbook.active
		return self.worksheet

	# get number of rows
	def get_max_row(self):
		max_row = self.worksheet.get_highest_row()
		# print('Rows:', max_row) #test
		return max_row

	# get number of columns
	def get_max_column(self):
		max_column = self.worksheet.get_highest_column()
		# print('Columns:', max_column) #test
		return max_column

	# get the letter label of last column ex 'A200'
	def get_sheet_max(self):
		letter = get_column_letter(self.get_max_column())
		# print('Last column:', letter) #test
		# get the last populated cell label
		sheet_max = letter + str(self.get_max_row())
		# print('Last cell:', sheet_max) #test
		return sheet_max

	# get column headers from excel file
	def get_column_headers(self):
		column_headers = []
		max_column = self.get_max_column()
		# from 1 to the last column (has to be +1)
		for column_number in range(1, max_column+1):
			# get the letter index of the current cell
			letter = get_column_letter(column_number)
			# append 1 to the letter to make it A1, B1, C1, etc
			cell = str(letter) + str(1)
			# append the cell index's value (A1 = formNo for example) to the column headers array
			column_headers.append(self.worksheet[cell].value)
			# print(worksheet[cell].value) #test
		# print('Column headers:', column_headers) #test
		return column_headers

	# we're going to create a dictionary to hold the contents of the rows in the excel file
	def get_blank_record(self):
		record = {}
		# create a dictionary key out of every column and set it equal to NULL
		# the end format will look like {formNo:NULL, firstname:NULL}
		column_headers = self.get_column_headers()
		for x in range(len(column_headers)):
			record[column_headers[x]] = 'NULL'
		# print("Dictionary:",record) #test
		return record

	# read the specified template into memory
	def get_template(self):
		blank_template = None
		with open(self.template, 'r') as file:
			blank_template = file.read()
		return blank_template
		
	def write_file(self, file_data, record):
		filename = self.get_filename(self.file_save_location, record, self.file_naming_convention, self.file_type)	
		fh = open(filename, "w")
		fh.write(file_data)
		fh.close()

	def append_file(self, file_data):
		with open(self.file_to_append, "a") as all:
			all.write(file_data)
		
	def populate_dictionary(self, record, workheet, row):
		for cellObj in row:
			# set the dictionary key to the contents of the cell
			# for example: set petitionerName to John Doe
			record[self.worksheet[str(cellObj.column)+'1'].value] = cellObj.value
		return record

	# construct a file name based on the user given paramters
	def get_filename(self, file_save_location, record, file_naming_convention, file_type):
		return str(file_save_location) + str(record[file_naming_convention]) + file_type

	# replace the targetted strings in the template with the corresponding excel file cells
	def fill_out_template(self, record, file_data):
		for x in record:
			string = '`' + str(x) + '`'
			file_data = file_data.replace(string, str(record[x]))
		return file_data

	# for each row in the excel sheet (start at A2 because A1 is just column headers)
	def transcribe(self):
		# get the last cell in the sheet
		sheet_max = self.get_sheet_max()
		# make the blank dictionary of column headers
		record = self.get_blank_record()
		# get the template to fill out
		blank_template = self.get_template()
		#Make one file or separate files?
		for rowOfCellObjects in self.worksheet['A2':sheet_max]: #sheet_max
			file_data = blank_template
			print("Populating dictionary\n")
			record = self.populate_dictionary(record, self.worksheet, rowOfCellObjects)
			print("Replacing template values\n")
			file_data = self.fill_out_template(record, file_data)
			print("Writing out edited file\n")
			self.write_file(file_data, record)
			self.append_file(file_data)

t = Transcriber("Excel Files/Oaths.xlsx", "Templates/Oath Template.txt", "Transcriptions/Oath Transcriptions/", "number", ".txt", "~All.doc")

t.transcribe()